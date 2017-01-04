from openpyxl import load_workbook
from openpyxl import Workbook

def get_cols(rows,value):
    for i,j in enumerate(rows):
        if j == value:
            return i
    raise ValueError("你所指定的列:\"{}\"不存在".format(value))

def getorcreate_sheet(wb,sheetname,head=None):
    if sheetname in wb.sheetnames:
        return wb[sheetname]
    else:
        ws = wb.create_sheet(sheetname)
        if head:
            ws.append(head)
        return ws

def split_sheet(ws,col_name):
    data = ws.values
    head = next(data)
    col_number = get_cols(head,col_name)
    new_wb = Workbook(write_only=True)

    for row in data:
         new_ws = getorcreate_sheet(new_wb,row[col_number],head)
         new_ws.append(row)
    return new_wb

if __name__ == "__main__":
    import sys
    try:
        filename = sys.argv[1]
        col_name = sys.argv[2]
    except Exception:
        print("用法：\n    python split_excel.py 文件名 分组列名称")
        sys.exit(1)
    wb = load_workbook(filename=filename)
    ws = wb.active
    wb2 = split_sheet(ws,col_name)
    wb2.save("new"+filename)
